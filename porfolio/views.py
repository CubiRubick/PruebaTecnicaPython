from django.shortcuts import render
from django.urls import reverse_lazy
from django.views.generic import TemplateView
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from .models import cliente
from django.db.models import Q
from openpyxl import Workbook
from django.http.response import HttpResponse

class porfolioDash(TemplateView):
    template_name = 'cliente_dash.html'

class porfolioSearch(TemplateView):
    template_name = 'client_search.html'

#lista de datos y buscador
def listar_cliente(request):
    busqueda = request.GET.get('buscar')
    if busqueda:
        clientes = cliente.objects.filter(Q(nombre__icontains=busqueda)|Q(email__icontains=busqueda)).distinct()
    else:
        clientes = cliente.objects.all()

    context = {
            'object_list':clientes
        }
    return render(request, 'porfolio/cliente_list.html', context)

#Ver datos unicos de un usuario y eliminacion de dichos datos
class porfolioDetail(DetailView, DeleteView):
    model = cliente
    success_url = reverse_lazy('porfolio:list')

#Creacion de nuevos usuarios
class porfolioCreated(CreateView):
    model = cliente
    success_url = reverse_lazy('porfolio:dash')
    fields = ['nombre', 'email', 'numeroT', 'descripcion']

#Actrualizacion de datos
class porfolioUpdate(UpdateView):
    model = cliente
    success_url = reverse_lazy('porfolio:list')
    fields = ['nombre', 'email', 'numeroT', 'descripcion']

#Nuestra clase hereda de la vista genérica TemplateView
class ReportePersonasExcel(TemplateView):
     
    #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        #Obtenemos todas las personas de nuestra base de datos
        personas = cliente.objects.all()
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'REPORTE DE PERSONAS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'nombre'
        ws['C3'] = 'correo electronico'
        ws['D3'] = 'numero de telefono'
        ws['E3'] = 'descripcion'       
        cont=4
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for persona in personas:
            ws.cell(row=cont,column=2).value = persona.nombre
            ws.cell(row=cont,column=3).value = persona.email
            ws.cell(row=cont,column=4).value = persona.numeroT
            ws.cell(row=cont,column=5).value = persona.descripcion
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReportePersonasExcel.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response